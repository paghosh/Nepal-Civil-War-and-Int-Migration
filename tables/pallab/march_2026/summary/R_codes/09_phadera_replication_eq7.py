"""
replicate_phadera.py

Replicates the empirical strategy from Phadera (2021) "Unfortunate Moms and
Unfortunate Children" using NLFS data with MIGRATION as the outcome variable
instead of height.

Key adaptation:
  - Original paper: Y = adult height (women), using NDHS 2016
  - This replication: Y = international migration / domestic migration, using NLFS
  - Same DID structure: conflict_v × cohort_c (Equation 7)
  - Same intergenerational framework: Equation 8

Outputs:
  1. Tables 1-3 style summary statistics
  2. Figure 4 (cohort diagram) and Figure 5 (exposure shares)
  3. Summary stats for male children in utero during conflict
  4. Table 4 style DID regression results (Eq 7) for migration outcomes
  5. All results by age groups: 0-5, 6-10, 11-18, 19-25, 26-40
"""

import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.formula.api as smf
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors as rl_colors
from reportlab.pdfgen import canvas
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import warnings, os, glob
warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════════════════════
BASE = "/Users/pallab.ghosh/Documents/GitHub/Nepal-Civil-War-and-Int-Migration"
DATA_PATH = f"{BASE}/data/Modified_Data/1_conflict_present_absentee_data.dta"
OUT_DIR = f"{BASE}/tables/pallab/march_2026/summary/results"
FIG_DIR = f"{BASE}/tables/pallab/march_2026/summary/figures"
os.makedirs(FIG_DIR, exist_ok=True)

SURVEY_YEAR = 2017
CONFLICT_START = 1996
CONFLICT_END = 2006

# ══════════════════════════════════════════════════════════════════════════════
# LOAD & PREPARE DATA
# ══════════════════════════════════════════════════════════════════════════════
print("Loading data...")
df = pd.read_stata(DATA_PATH, convert_categoricals=False)

df['sex_num'] = pd.to_numeric(df['sex'], errors='coerce')
df['male'] = (df['sex_num'] == 1).astype(float)
df['female'] = (df['sex_num'] == 2).astype(float)
df['age_num'] = pd.to_numeric(df['age'], errors='coerce')
df['birth_year'] = SURVEY_YEAR - df['age_num']
df['age_at_conflict_start'] = CONFLICT_START - df['birth_year']
df['age_at_conflict_end'] = CONFLICT_END - df['birth_year']

# Casualty variables as numeric + fill for absentees
for cv in ['cas_own_any','cas_own_fatal','mwar_own_any','mwar_own_fatal',
           'cas_nbr_any','cas_nbr_fatal','mwar_nbr_any','mwar_nbr_fatal']:
    df[cv] = pd.to_numeric(df[cv], errors='coerce')

dist_cas = df.dropna(subset=['cas_own_any']).groupby('dist')[
    ['cas_own_any','cas_own_fatal','mwar_own_any','mwar_own_fatal',
     'cas_nbr_any','cas_nbr_fatal','mwar_nbr_any','mwar_nbr_fatal']].first()
for cv in dist_cas.columns:
    missing = df[cv].isna()
    df.loc[missing, cv] = df.loc[missing, 'dist'].map(dist_cas[cv])

# Ethnicity
caste = pd.to_numeric(df['caste'], errors='coerce')
df['caste_num'] = caste
df['caste_num'] = df.groupby(['psu','hhld'])['caste_num'].transform(lambda x: x.ffill().bfill())
hill_high={1,2,14,20,27,48,49}
hill_jan={3,5,6,10,11,13,24,29,32,36,45,46,60,61,62,66,67,69,74,77,78,79,80,81,89,90,91,92,94,97,98,100,110,119,120,121,124,125,126,127,130,131,132,133,134,135,136,137,138,139,140,992}
dalit_set={8,12,15,17,22,23,25,38,39,40,41,50,70,75,76,83,93,991}
def get_eth(c):
    if pd.isna(c): return None
    c=int(c)
    if c in hill_high: return 'Hill High Caste'
    if c in hill_jan: return 'Hill Janajati'
    if c in dalit_set: return 'Dalit'
    if c==7: return 'Muslim'
    return 'Other'
df['ethnicity'] = df['caste_num'].apply(get_eth)
df['high_caste'] = (df['ethnicity'] == 'Hill High Caste').astype(float)
df.loc[df['ethnicity'].isna(), 'high_caste'] = np.nan

# Education
gc = pd.to_numeric(df['grade_comp'], errors='coerce')
def get_edu(g):
    if pd.isna(g): return None
    if g in [16,17]: return 'No Education'
    if 0<=g<=5: return 'Primary (1-5)'
    if 6<=g<=12: return 'Secondary (6-12)'
    if g>=13: return 'Tertiary'
    return None
df['edu_cat'] = gc.apply(get_edu)
df['years_education'] = gc.where(~gc.isin([16,17]), 0)

# Migration outcomes (binary)
df['intl_mig'] = df['international_absentee_only'].copy()
df['dom_mig'] = df['national'].copy()

# Cohort assignment (following Phadera's Figure 4)
# Treatment cohorts (exposed during critical growth periods):
#   Age 0-3 at conflict start → born 1993-1996
#   Age 4-8 at conflict start → born 1988-1992
#   Age 9-15 at conflict start → born 1981-1987
# Control cohort:
#   Age 16-21 at conflict start → born 1975-1980 (gained full height before war)
# Placebo:
#   Age 22-29 at conflict start → born 1967-1974
def assign_cohort(age_cs):
    if pd.isna(age_cs): return None
    if age_cs < 0: return 'In Utero/Post-War'
    if age_cs <= 3: return 'Age 0-3'
    if age_cs <= 8: return 'Age 4-8'
    if age_cs <= 15: return 'Age 9-15'
    if age_cs <= 21: return 'Age 16-21 (Control)'
    if age_cs <= 29: return 'Age 22-29 (Placebo)'
    return 'Age 30+'

df['cohort'] = df['age_at_conflict_start'].apply(assign_cohort)

# Our age groups of interest
def assign_age_group(age_cs):
    if pd.isna(age_cs): return None
    if age_cs < 0: return None
    if age_cs <= 5: return '0-5'
    if age_cs <= 10: return '6-10'
    if age_cs <= 18: return '11-18'
    if age_cs <= 25: return '19-25'
    if age_cs <= 40: return '26-40'
    return None
df['age_group'] = df['age_at_conflict_start'].apply(assign_age_group)

# In utero at conflict: born during Feb 1996 - Nov 1996
# (conflict started Feb 1996, gestation ~9 months)
# age_at_conflict_start would be ~0 or negative (born 1996-1997)
# More precisely: in utero = born 1996 or early 1997
df['in_utero'] = ((df['birth_year'] >= 1996) & (df['birth_year'] <= 1997)).astype(float)

# Treatment based on conflict median (as in our existing tables)
for cv in ['mwar_own_any', 'mwar_own_fatal', 'cas_own_any', 'cas_own_fatal']:
    med = df[cv].median()
    df[f'treat_{cv}'] = (df[cv] > med).astype(float)
    df.loc[df[cv].isna(), f'treat_{cv}'] = np.nan

# District fixed effects
df['dist_num'] = pd.to_numeric(df['dist'], errors='coerce')

# PSU (village) identifier
df['psu_num'] = pd.to_numeric(df['psu'], errors='coerce')

print(f"Data loaded: {len(df)} rows, {len(df.columns)} columns")
print(f"Age at conflict start range: {df['age_at_conflict_start'].min():.0f} to {df['age_at_conflict_start'].max():.0f}")

# ══════════════════════════════════════════════════════════════════════════════
# TABLE 1: SUMMARY STATISTICS - ALL INDIVIDUALS (like Phadera Table 2)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  GENERATING TABLE 1: Summary Statistics (All Individuals)")
print("="*70)

def compute_summary_row(sub_all, sub_treat, sub_ctrl, var, label):
    """Compute mean, sd, and diff with t-test for a variable."""
    all_vals = sub_all[var].dropna()
    t_vals = sub_treat[var].dropna()
    c_vals = sub_ctrl[var].dropna()

    all_mean = all_vals.mean() if len(all_vals) > 0 else np.nan
    all_sd = all_vals.std() if len(all_vals) > 0 else np.nan
    t_mean = t_vals.mean() if len(t_vals) > 0 else np.nan
    t_sd = t_vals.std() if len(t_vals) > 0 else np.nan
    c_mean = c_vals.mean() if len(c_vals) > 0 else np.nan
    c_sd = c_vals.std() if len(c_vals) > 0 else np.nan

    diff = t_mean - c_mean if not (np.isnan(t_mean) or np.isnan(c_mean)) else np.nan

    # t-test
    stars = ''
    se = np.nan
    if len(t_vals) >= 2 and len(c_vals) >= 2:
        _, p = stats.ttest_ind(t_vals, c_vals, equal_var=False)
        se = np.sqrt(t_vals.var()/len(t_vals) + c_vals.var()/len(c_vals))
        if p < 0.01: stars = '***'
        elif p < 0.05: stars = '**'
        elif p < 0.10: stars = '*'

    return {
        'label': label,
        'all_mean': all_mean, 'all_sd': all_sd,
        't_mean': t_mean, 't_sd': t_sd,
        'c_mean': c_mean, 'c_sd': c_sd,
        'diff': diff, 'se': se, 'stars': stars
    }

# Define treatment/control as in Phadera:
# Treatment = Age 0-15 at conflict start (exposed during critical growth)
# Control = Age 16-29 at conflict start (already grown)
# We adapt for our age groups too

# Full sample for Table 1: all with valid age
sample = df[df['age_at_conflict_start'].between(0, 40)].copy()
treat = sample[sample['age_at_conflict_start'].between(0, 17)]
ctrl = sample[sample['age_at_conflict_start'].between(18, 40)]

table1_rows = []
# Panel A: Outcomes (Migration)
table1_rows.append({'label': 'Panel A: Outcomes', 'section': True})
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'intl_mig', 'International Migration (=1)'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'dom_mig', 'Domestic Migration (=1)'))

# Panel B: Exposure to conflict
table1_rows.append({'label': 'Panel B: Exposure to Conflict', 'section': True})
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'mwar_own_any', 'Months of War (own village)'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'cas_own_any', 'Number of Casualties (own village)'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'mwar_nbr_any', 'Months of War (incl. neighbors)'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'cas_nbr_any', 'Casualties (incl. neighbors)'))

# Panel C: Controls
table1_rows.append({'label': 'Panel C: Controls', 'section': True})
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'age_num', 'Current Age'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'male', 'Male (=1)'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'high_caste', 'High Caste (=1)'))
table1_rows.append(compute_summary_row(sample, treat, ctrl, 'years_education', 'Years of Education'))

# Add N
table1_rows.append({'label': f'Number of Individuals', 'section': False,
                     'all_mean': len(sample), 'all_sd': np.nan,
                     't_mean': len(treat), 't_sd': np.nan,
                     'c_mean': len(ctrl), 'c_sd': np.nan,
                     'diff': np.nan, 'se': np.nan, 'stars': ''})

# ══════════════════════════════════════════════════════════════════════════════
# TABLE 2: Summary by Migration Type (like Phadera Table 3)
# ══════════════════════════════════════════════════════════════════════════════
print("  GENERATING TABLE 2: Summary Statistics by Migration Type")

intl_sample = sample[sample['intl_mig'] == 1]
dom_sample = sample[sample['dom_mig'] == 1]
baseline_sample = sample[(sample['intl_mig'] != 1) & (sample['dom_mig'] != 1)]

table2_rows = []
for var, label in [('age_num', 'Current Age'), ('male', 'Male (=1)'),
                    ('high_caste', 'High Caste (=1)'), ('years_education', 'Years of Education'),
                    ('mwar_own_any', 'Months of War'), ('cas_own_any', 'Number of Casualties')]:
    for sub, sub_name in [(baseline_sample, 'Baseline'), (intl_sample, 'Intl Migration'),
                           (dom_sample, 'Dom Migration')]:
        vals = sub[var].dropna()
        table2_rows.append({
            'label': label, 'group': sub_name,
            'mean': vals.mean() if len(vals) > 0 else np.nan,
            'sd': vals.std() if len(vals) > 0 else np.nan,
            'n': len(vals)
        })

# ══════════════════════════════════════════════════════════════════════════════
# FIGURE 4: Cohort Diagram (Age at Conflict Start)
# ══════════════════════════════════════════════════════════════════════════════
print("  GENERATING FIGURE 4: Cohort Diagram")

fig, ax = plt.subplots(figsize=(14, 5))

cohort_colors = {
    'In Utero': '#FFB3BA',
    'Age 0-5': '#FF6B6B',
    'Age 6-10': '#FF8E53',
    'Age 11-18': '#FFC107',
    'Age 19-25': '#4CAF50',
    'Age 26-40': '#2196F3',
}

# Draw bars for each age group
age_groups_fig = [
    ('In Utero\n(born after 1996)', -3, 0, '#FFB3BA'),
    ('Age 0-5\nat conflict start', 0, 5, '#FF6B6B'),
    ('Age 6-10\nat conflict start', 6, 10, '#FF8E53'),
    ('Age 11-18\nat conflict start', 11, 18, '#FFC107'),
    ('Age 19-25\nat conflict start\n(Control 1)', 19, 25, '#4CAF50'),
    ('Age 26-40\nat conflict start\n(Control 2)', 26, 40, '#2196F3'),
]

y_pos = 0
for label, lo, hi, color in age_groups_fig:
    width = hi - lo + 1
    ax.barh(y_pos, width, left=lo, height=0.6, color=color, edgecolor='black', linewidth=0.5)
    ax.text(lo + width/2, y_pos, label, ha='center', va='center', fontsize=7, fontweight='bold')
    y_pos += 1

ax.set_xlabel('Age at Start of Civil War (February 1996)', fontsize=11, fontweight='bold')
ax.set_title('Figure 4: Cohorts by Age at the Start of the War in 1996\n(Adapted from Phadera, 2021)',
             fontsize=13, fontweight='bold')
ax.set_yticks([])
ax.axvline(x=0, color='red', linestyle='--', linewidth=1, alpha=0.7, label='Conflict Start (1996)')
ax.set_xlim(-5, 42)
ax.legend(loc='upper right')
plt.tight_layout()
plt.savefig(f'{FIG_DIR}/figure_4_cohort_diagram.png', dpi=300, bbox_inches='tight')
plt.savefig(f'{FIG_DIR}/figure_4_cohort_diagram.pdf', bbox_inches='tight')
plt.close()
print(f"  Saved: {FIG_DIR}/figure_4_cohort_diagram.png/pdf")

# ══════════════════════════════════════════════════════════════════════════════
# FIGURE 5: Migration Rate by Age Group and Conflict Intensity
# ══════════════════════════════════════════════════════════════════════════════
print("  GENERATING FIGURE 5: Migration Rates by Conflict Intensity")

fig, axes = plt.subplots(1, 2, figsize=(16, 7))

age_bins = [(0, 5), (6, 10), (11, 18), (19, 25), (26, 40)]
age_labels = ['0-5', '6-10', '11-18', '19-25', '26-40']

med_mwar = df['mwar_own_any'].median()

for ax_idx, (outcome, outcome_label) in enumerate([('intl_mig', 'International Migration'),
                                                      ('dom_mig', 'Domestic Migration')]):
    ax = axes[ax_idx]

    treat_rates = []
    ctrl_rates = []

    for lo, hi in age_bins:
        sub = df[df['age_at_conflict_start'].between(lo, hi)]
        t = sub[sub['mwar_own_any'] > med_mwar]
        c = sub[sub['mwar_own_any'] <= med_mwar]

        t_rate = t[outcome].mean() * 100 if len(t) > 0 else 0
        c_rate = c[outcome].mean() * 100 if len(c) > 0 else 0
        treat_rates.append(t_rate)
        ctrl_rates.append(c_rate)

    x = np.arange(len(age_labels))
    width = 0.35

    bars1 = ax.bar(x - width/2, treat_rates, width, label='High Conflict (Treatment)',
                    color='#E74C3C', edgecolor='black', linewidth=0.5)
    bars2 = ax.bar(x + width/2, ctrl_rates, width, label='Low Conflict (Control)',
                    color='#3498DB', edgecolor='black', linewidth=0.5)

    ax.set_xlabel('Age at Conflict Start (1996)', fontsize=11, fontweight='bold')
    ax.set_ylabel(f'{outcome_label} Rate (%)', fontsize=11, fontweight='bold')
    ax.set_title(f'{outcome_label} by Age Cohort\nand Conflict Intensity', fontsize=12, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(age_labels)
    ax.legend()
    ax.grid(axis='y', alpha=0.3)

    # Add value labels
    for bar in bars1:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., h + 0.2, f'{h:.1f}%',
                ha='center', va='bottom', fontsize=8)
    for bar in bars2:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., h + 0.2, f'{h:.1f}%',
                ha='center', va='bottom', fontsize=8)

plt.suptitle('Figure 5: Migration Rates by Age at Conflict Start and Conflict Intensity\n'
             f'(Treatment: Months of War > median ({med_mwar:.0f}); Control: ≤ median)',
             fontsize=14, fontweight='bold', y=1.02)
plt.tight_layout()
plt.savefig(f'{FIG_DIR}/figure_5_migration_by_conflict.png', dpi=300, bbox_inches='tight')
plt.savefig(f'{FIG_DIR}/figure_5_migration_by_conflict.pdf', bbox_inches='tight')
plt.close()
print(f"  Saved: {FIG_DIR}/figure_5_migration_by_conflict.png/pdf")

# ══════════════════════════════════════════════════════════════════════════════
# TABLE 3: Summary Stats for Male Children In Utero During Conflict
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  GENERATING TABLE 3: Summary Stats - Males In Utero During Conflict")
print("="*70)

# In utero during conflict: born between 1996 and 2006 (conflict period)
# Males only
males_in_utero = df[(df['male'] == 1) &
                     (df['birth_year'] >= 1996) &
                     (df['birth_year'] <= 2006)].copy()

# Split by conflict intensity
med_cas = males_in_utero['cas_own_any'].median()
m_treat = males_in_utero[males_in_utero['cas_own_any'] > med_cas]
m_ctrl = males_in_utero[males_in_utero['cas_own_any'] <= med_cas]

table3_rows = []
table3_rows.append({'label': 'Panel A: Outcomes', 'section': True})
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'intl_mig', 'International Migration (=1)'))
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'dom_mig', 'Domestic Migration (=1)'))

table3_rows.append({'label': 'Panel B: Exposure to Conflict', 'section': True})
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'mwar_own_any', 'Months of War'))
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'cas_own_any', 'Number of Casualties'))

table3_rows.append({'label': 'Panel C: Demographics', 'section': True})
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'age_num', 'Current Age'))
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'high_caste', 'High Caste (=1)'))
table3_rows.append(compute_summary_row(males_in_utero, m_treat, m_ctrl, 'years_education', 'Years of Education'))

table3_rows.append({'label': f'Number of Males', 'section': False,
                     'all_mean': len(males_in_utero), 'all_sd': np.nan,
                     't_mean': len(m_treat), 't_sd': np.nan,
                     'c_mean': len(m_ctrl), 'c_sd': np.nan,
                     'diff': np.nan, 'se': np.nan, 'stars': ''})

print(f"  Males in utero during conflict: {len(males_in_utero)}")
print(f"  Treatment (high casualty): {len(m_treat)}, Control: {len(m_ctrl)}")

# ══════════════════════════════════════════════════════════════════════════════
# TABLE 4: DID REGRESSION - Eq. (7) with Migration Outcomes
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  GENERATING TABLE 4: DID Regressions (Equation 7)")
print("="*70)

# Following Phadera Eq 7:
# Y_i = sum_c beta_c * (conflict_v × cohort_c) + cohort_FE + birth_year_FE + district_FE + controls + e
#
# We use district FE (since we don't have village-level GPS in NLFS)
# Conflict variable: months of war in own district
# Cohorts: Age 0-3, 4-8, 9-15 (treatment), 16-21 (control), 22-29 (placebo)

# Prepare regression sample
reg_sample = df[df['age_at_conflict_start'].between(0, 40) &
                df['mwar_own_any'].notna() &
                df['age_num'].notna()].copy()

# Create cohort dummies
reg_sample['cohort_0_3'] = (reg_sample['age_at_conflict_start'].between(0, 3)).astype(float)
reg_sample['cohort_4_8'] = (reg_sample['age_at_conflict_start'].between(4, 8)).astype(float)
reg_sample['cohort_9_15'] = (reg_sample['age_at_conflict_start'].between(9, 15)).astype(float)
reg_sample['cohort_16_21'] = (reg_sample['age_at_conflict_start'].between(16, 21)).astype(float)  # control
reg_sample['cohort_22_29'] = (reg_sample['age_at_conflict_start'].between(22, 29)).astype(float)  # placebo

# Our age groups
reg_sample['cohort_0_5'] = (reg_sample['age_at_conflict_start'].between(0, 5)).astype(float)
reg_sample['cohort_6_10'] = (reg_sample['age_at_conflict_start'].between(6, 10)).astype(float)
reg_sample['cohort_11_18'] = (reg_sample['age_at_conflict_start'].between(11, 18)).astype(float)
reg_sample['cohort_19_25'] = (reg_sample['age_at_conflict_start'].between(19, 25)).astype(float)  # control
reg_sample['cohort_26_40'] = (reg_sample['age_at_conflict_start'].between(26, 40)).astype(float)  # placebo

# Interaction terms: conflict × cohort
for conflict_var in ['mwar_own_any', 'cas_own_any']:
    for cohort_name in ['0_3', '4_8', '9_15', '22_29',
                         '0_5', '6_10', '11_18', '26_40']:
        reg_sample[f'{conflict_var}_x_{cohort_name}'] = (
            reg_sample[conflict_var] * reg_sample[f'cohort_{cohort_name}']
        )

# Fill high_caste NaN
reg_sample['high_caste'] = reg_sample['high_caste'].fillna(0)

# Run regressions
results_table4 = {}

for outcome in ['intl_mig', 'dom_mig']:
    for conflict_var in ['mwar_own_any', 'cas_own_any']:
        # Phadera-style cohorts (Table 4 replication)
        interactions = [f'{conflict_var}_x_22_29', f'{conflict_var}_x_9_15',
                        f'{conflict_var}_x_4_8', f'{conflict_var}_x_0_3']

        # Build formula
        # District FE via C(dist_num) - absorb district effects
        formula = f'{outcome} ~ ' + ' + '.join(interactions)
        formula += ' + cohort_0_3 + cohort_4_8 + cohort_9_15 + cohort_22_29'
        formula += ' + high_caste + C(dist_num)'

        try:
            valid = reg_sample.dropna(subset=[outcome] + interactions + ['high_caste', 'dist_num'])
            model = smf.ols(formula, data=valid).fit(cov_type='cluster',
                                                       cov_kwds={'groups': valid['psu_num']})

            key = f'{outcome}_{conflict_var}_phadera'
            results_table4[key] = {}
            for interaction in interactions:
                coef = model.params.get(interaction, np.nan)
                se = model.bse.get(interaction, np.nan)
                pval = model.pvalues.get(interaction, np.nan)
                stars = '***' if pval < 0.01 else ('**' if pval < 0.05 else ('*' if pval < 0.10 else ''))
                results_table4[key][interaction] = {
                    'coef': coef, 'se': se, 'pval': pval, 'stars': stars
                }
            results_table4[key]['nobs'] = model.nobs
            results_table4[key]['r2_adj'] = model.rsquared_adj

            # Control mean
            ctrl_mask = valid['cohort_16_21'] == 1
            ctrl_mean = valid.loc[ctrl_mask, outcome].mean() if ctrl_mask.any() else np.nan
            results_table4[key]['ctrl_mean'] = ctrl_mean

            print(f"  {outcome} ~ {conflict_var} (Phadera cohorts): N={model.nobs:.0f}, R2={model.rsquared_adj:.4f}")
        except Exception as e:
            print(f"  ERROR: {outcome} ~ {conflict_var}: {e}")

        # Our age group cohorts
        interactions2 = [f'{conflict_var}_x_26_40', f'{conflict_var}_x_11_18',
                         f'{conflict_var}_x_6_10', f'{conflict_var}_x_0_5']

        formula2 = f'{outcome} ~ ' + ' + '.join(interactions2)
        formula2 += ' + cohort_0_5 + cohort_6_10 + cohort_11_18 + cohort_26_40'
        formula2 += ' + high_caste + C(dist_num)'

        try:
            valid2 = reg_sample.dropna(subset=[outcome] + interactions2 + ['high_caste', 'dist_num'])
            model2 = smf.ols(formula2, data=valid2).fit(cov_type='cluster',
                                                          cov_kwds={'groups': valid2['psu_num']})

            key2 = f'{outcome}_{conflict_var}_agegroups'
            results_table4[key2] = {}
            for interaction in interactions2:
                coef = model2.params.get(interaction, np.nan)
                se = model2.bse.get(interaction, np.nan)
                pval = model2.pvalues.get(interaction, np.nan)
                stars = '***' if pval < 0.01 else ('**' if pval < 0.05 else ('*' if pval < 0.10 else ''))
                results_table4[key2][interaction] = {
                    'coef': coef, 'se': se, 'pval': pval, 'stars': stars
                }
            results_table4[key2]['nobs'] = model2.nobs
            results_table4[key2]['r2_adj'] = model2.rsquared_adj

            ctrl_mask2 = valid2['cohort_19_25'] == 1
            ctrl_mean2 = valid2.loc[ctrl_mask2, outcome].mean() if ctrl_mask2.any() else np.nan
            results_table4[key2]['ctrl_mean'] = ctrl_mean2

            print(f"  {outcome} ~ {conflict_var} (Our age groups): N={model2.nobs:.0f}, R2={model2.rsquared_adj:.4f}")
        except Exception as e:
            print(f"  ERROR: {outcome} ~ {conflict_var} (age groups): {e}")

# ══════════════════════════════════════════════════════════════════════════════
# WRITE ALL TABLES TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  WRITING EXCEL OUTPUT")
print("="*70)

wb = Workbook()

# ── Style constants ──
NAVY = "1B2A4A"
TEAL = "2E8B8B"
GOLD = "D4A843"
WHITE = "FFFFFF"
LGRAY = "F5F5F5"
MGRAY = "E0E0E0"
LTEAL = "E0F2F1"

hdr_font = Font(name='Arial', size=10, bold=True, color=WHITE)
hdr_fill = PatternFill('solid', fgColor=NAVY)
sec_font = Font(name='Arial', size=10, bold=True, color=TEAL)
sec_fill = PatternFill('solid', fgColor=LTEAL)
data_font = Font(name='Arial', size=9, color='1A1A1A')
note_font = Font(name='Arial', size=7.5, italic=True, color='666666')
thin_border = Border(bottom=Side(style='thin', color=MGRAY))

def fmt(val, decimals=2):
    if pd.isna(val) or val is None: return ''
    if isinstance(val, int) or (isinstance(val, float) and val == int(val) and abs(val) > 100):
        return f'{int(val):,}'
    return f'{val:.{decimals}f}'

def write_summary_table(ws, title, rows, col_headers):
    """Write a summary statistics table to a worksheet."""
    r = 1
    COLS = len(col_headers) + 1

    # Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(name='Arial', size=12, bold=True, color=NAVY)
    r += 2

    # Headers
    for ci, hdr in enumerate(col_headers, start=2):
        c = ws.cell(row=r, column=ci, value=hdr)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.cell(row=r, column=1).fill = hdr_fill
    r += 1

    # Data rows
    alt = False
    for row_data in rows:
        if row_data.get('section'):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
            c = ws.cell(row=r, column=1, value=row_data['label'])
            c.font = sec_font
            c.fill = sec_fill
            for ci in range(2, COLS+1):
                ws.cell(row=r, column=ci).fill = sec_fill
            r += 1
            alt = False
            continue

        bg = PatternFill('solid', fgColor=LGRAY) if alt else PatternFill('solid', fgColor=WHITE)

        # Label
        c = ws.cell(row=r, column=1, value=row_data['label'])
        c.font = data_font
        c.fill = bg
        c.alignment = Alignment(indent=1)

        # Values
        vals = [
            fmt(row_data.get('all_mean')), f"[{fmt(row_data.get('all_sd'))}]",
            fmt(row_data.get('t_mean')), f"[{fmt(row_data.get('t_sd'))}]",
            fmt(row_data.get('c_mean')), f"[{fmt(row_data.get('c_sd'))}]",
            f"{fmt(row_data.get('diff'))}{row_data.get('stars','')}" if not pd.isna(row_data.get('diff', np.nan)) else '',
            f"({fmt(row_data.get('se'))})" if not pd.isna(row_data.get('se', np.nan)) else ''
        ]

        for ci, val in enumerate(vals, start=2):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = data_font
            c.fill = bg
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        r += 1
        alt = not alt

    # Set column widths
    ws.column_dimensions['A'].width = 28
    for ci in range(2, COLS+1):
        ws.column_dimensions[chr(64+ci)].width = 14

    return r

# Sheet 1: Table 1 - Summary Statistics
ws1 = wb.active
ws1.title = "Table 1 - Summary Stats"
col_hdrs_1 = ['All\nMean', 'All\n[SD]', 'Treatment\n(Age 0-17)\nMean', 'Treatment\n[SD]',
              'Control\n(Age 18-40)\nMean', 'Control\n[SD]', 'Difference', '(SE)']
r = write_summary_table(ws1,
    'Table 1: Summary Statistics of Individuals (Age 0-40 at Conflict Start)',
    table1_rows, col_hdrs_1)
# Notes
ws1.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=9)
ws1.cell(row=r+1, column=1, value="Note: Treatment = aged 0-17 at conflict start. Control = aged 18-40. Standard deviations in brackets, standard errors in parentheses.").font = note_font
ws1.cell(row=r+2, column=1, value="*** p<0.01, ** p<0.05, * p<0.10. Data: Nepal Labor Force Survey, conflict data from INSEC.").font = note_font

# Sheet 2: Table 3 - Males In Utero
ws3 = wb.create_sheet("Table 3 - Males In Utero")
col_hdrs_3 = ['All\nMean', 'All\n[SD]', 'High Conflict\nMean', 'Treatment\n[SD]',
              'Low Conflict\nMean', 'Control\n[SD]', 'Difference', '(SE)']
r3 = write_summary_table(ws3,
    'Table 3: Summary Statistics - Male Children In Utero During Conflict (Born 1996-2006)',
    table3_rows, col_hdrs_3)
ws3.cell(row=r3+1, column=1, value="Note: Males born 1996-2006 (during conflict). Treatment = above-median casualties. Standard deviations in brackets, SE in parentheses.").font = note_font
ws3.cell(row=r3+2, column=1, value="*** p<0.01, ** p<0.05, * p<0.10").font = note_font

# Sheet 3: Table 4 - DID Regression Results (Phadera cohorts)
ws4 = wb.create_sheet("Table 4 - DID Phadera Cohorts")
r4 = 1
ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=9)
ws4.cell(row=r4, column=1, value='Table 4: Impact on Migration by Age at Start of the Civil War (Equation 7)').font = Font(name='Arial', size=12, bold=True, color=NAVY)
r4 += 2

# Column headers
outcome_labels = {'intl_mig': 'International Migration', 'dom_mig': 'Domestic Migration'}
conflict_labels = {'mwar_own_any': 'Months of War', 'cas_own_any': 'Casualty Count'}

col = 2
for outcome in ['intl_mig', 'dom_mig']:
    ws4.merge_cells(start_row=r4, start_column=col, end_row=r4, end_column=col+1)
    c = ws4.cell(row=r4, column=col, value=outcome_labels[outcome])
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
    ws4.cell(row=r4, column=col+1).fill = hdr_fill
    col += 2
r4 += 1

col = 2
for outcome in ['intl_mig', 'dom_mig']:
    for conf_var in ['mwar_own_any', 'cas_own_any']:
        c = ws4.cell(row=r4, column=col, value=f'({col-1})\n{conflict_labels[conf_var]}')
        c.font = Font(name='Arial', size=8, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = Border(bottom=Side(style='medium', color=NAVY))
        col += 1
r4 += 1

# Regression rows
cohort_interactions = [
    ('22_29', 'Age 22-29 × Conflict'),
    ('9_15', 'Age 9-15 × Conflict'),
    ('4_8', 'Age 4-8 × Conflict'),
    ('0_3', 'Age 0-3 × Conflict'),
]

for cohort_suffix, row_label in cohort_interactions:
    # Coefficient row
    ws4.cell(row=r4, column=1, value=row_label).font = Font(name='Arial', size=9, bold=True)
    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_phadera'
            inter_name = f'{conf_var}_x_{cohort_suffix}'
            if key in results_table4 and inter_name in results_table4[key]:
                res = results_table4[key][inter_name]
                val_str = f"{res['coef']:.6f}{res['stars']}"
                ws4.cell(row=r4, column=col, value=val_str).font = data_font
                ws4.cell(row=r4, column=col).alignment = Alignment(horizontal='center')
            col += 1
    r4 += 1

    # SE row
    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_phadera'
            inter_name = f'{conf_var}_x_{cohort_suffix}'
            if key in results_table4 and inter_name in results_table4[key]:
                res = results_table4[key][inter_name]
                ws4.cell(row=r4, column=col, value=f"({res['se']:.6f})").font = Font(name='Arial', size=8, color='666666')
                ws4.cell(row=r4, column=col).alignment = Alignment(horizontal='center')
            col += 1
    r4 += 1

# Bottom stats
r4 += 1
for stat_label, stat_key in [('Observations', 'nobs'), ('Adjusted R-squared', 'r2_adj'), ('Control Mean (Age 16-21)', 'ctrl_mean')]:
    ws4.cell(row=r4, column=1, value=stat_label).font = Font(name='Arial', size=9, bold=True)
    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_phadera'
            if key in results_table4:
                val = results_table4[key].get(stat_key, np.nan)
                if stat_key == 'nobs':
                    ws4.cell(row=r4, column=col, value=f"{int(val):,}" if not np.isnan(val) else '').font = data_font
                else:
                    ws4.cell(row=r4, column=col, value=f"{val:.4f}" if not np.isnan(val) else '').font = data_font
                ws4.cell(row=r4, column=col).alignment = Alignment(horizontal='center')
            col += 1
    r4 += 1

ws4.cell(row=r4+1, column=1, value="Controls: District FE, cohort dummies, high caste indicator").font = note_font
ws4.cell(row=r4+2, column=1, value="Standard errors clustered at PSU level in parentheses. *** p<0.01, ** p<0.05, * p<0.10").font = note_font
ws4.column_dimensions['A'].width = 28
for ci in range(2, 7):
    ws4.column_dimensions[chr(64+ci)].width = 18

# Sheet 4: Table 5 - DID with our age groups
ws5 = wb.create_sheet("Table 5 - DID Age Groups")
r5 = 1
ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=9)
ws5.cell(row=r5, column=1, value='Table 5: Impact on Migration by Age Group at Start of the Civil War').font = Font(name='Arial', size=12, bold=True, color=NAVY)
r5 += 2

col = 2
for outcome in ['intl_mig', 'dom_mig']:
    ws5.merge_cells(start_row=r5, start_column=col, end_row=r5, end_column=col+1)
    c = ws5.cell(row=r5, column=col, value=outcome_labels[outcome])
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
    ws5.cell(row=r5, column=col+1).fill = hdr_fill
    col += 2
r5 += 1

col = 2
for outcome in ['intl_mig', 'dom_mig']:
    for conf_var in ['mwar_own_any', 'cas_own_any']:
        c = ws5.cell(row=r5, column=col, value=f'({col-1})\n{conflict_labels[conf_var]}')
        c.font = Font(name='Arial', size=8, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = Border(bottom=Side(style='medium', color=NAVY))
        col += 1
r5 += 1

cohort_interactions2 = [
    ('26_40', 'Age 26-40 × Conflict (Placebo)'),
    ('11_18', 'Age 11-18 × Conflict'),
    ('6_10', 'Age 6-10 × Conflict'),
    ('0_5', 'Age 0-5 × Conflict'),
]

for cohort_suffix, row_label in cohort_interactions2:
    ws5.cell(row=r5, column=1, value=row_label).font = Font(name='Arial', size=9, bold=True)
    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_agegroups'
            inter_name = f'{conf_var}_x_{cohort_suffix}'
            if key in results_table4 and inter_name in results_table4[key]:
                res = results_table4[key][inter_name]
                ws5.cell(row=r5, column=col, value=f"{res['coef']:.6f}{res['stars']}").font = data_font
                ws5.cell(row=r5, column=col).alignment = Alignment(horizontal='center')
            col += 1
    r5 += 1

    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_agegroups'
            inter_name = f'{conf_var}_x_{cohort_suffix}'
            if key in results_table4 and inter_name in results_table4[key]:
                res = results_table4[key][inter_name]
                ws5.cell(row=r5, column=col, value=f"({res['se']:.6f})").font = Font(name='Arial', size=8, color='666666')
                ws5.cell(row=r5, column=col).alignment = Alignment(horizontal='center')
            col += 1
    r5 += 1

r5 += 1
for stat_label, stat_key in [('Observations', 'nobs'), ('Adjusted R-squared', 'r2_adj'), ('Control Mean (Age 19-25)', 'ctrl_mean')]:
    ws5.cell(row=r5, column=1, value=stat_label).font = Font(name='Arial', size=9, bold=True)
    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_agegroups'
            if key in results_table4:
                val = results_table4[key].get(stat_key, np.nan)
                if stat_key == 'nobs':
                    ws5.cell(row=r5, column=col, value=f"{int(val):,}" if not np.isnan(val) else '').font = data_font
                else:
                    ws5.cell(row=r5, column=col, value=f"{val:.4f}" if not np.isnan(val) else '').font = data_font
                ws5.cell(row=r5, column=col).alignment = Alignment(horizontal='center')
            col += 1
    r5 += 1

ws5.cell(row=r5+1, column=1, value="Controls: District FE, cohort dummies, high caste indicator. Control group: Age 19-25.").font = note_font
ws5.cell(row=r5+2, column=1, value="Standard errors clustered at PSU level in parentheses. *** p<0.01, ** p<0.05, * p<0.10").font = note_font
ws5.column_dimensions['A'].width = 32
for ci in range(2, 7):
    ws5.column_dimensions[chr(64+ci)].width = 18

# Save workbook
xlsx_path = f"{OUT_DIR}/phadera_replication_tables.xlsx"
wb.save(xlsx_path)
print(f"\n  XLSX saved: {xlsx_path}")

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE LaTeX FILE
# ══════════════════════════════════════════════════════════════════════════════
print("  GENERATING LaTeX output...")

def write_tex_summary_table(rows, title, label, col_headers, note):
    """Generate LaTeX code for a summary table."""
    lines = []
    lines.append("\\begin{table}[htbp]")
    lines.append("\\centering")
    lines.append("\\footnotesize")
    lines.append(f"\\caption{{{title}}}")
    lines.append(f"\\label{{{label}}}")
    ncols = len(col_headers) + 1
    col_spec = 'l' + 'r' * len(col_headers)
    lines.append(f"\\begin{{tabular}}{{{col_spec}}}")
    lines.append("\\toprule")
    lines.append(" & " + " & ".join(col_headers) + " \\\\")
    lines.append("\\midrule")

    for row_data in rows:
        if row_data.get('section'):
            lines.append(f"\\textit{{{row_data['label']}}} " + "& " * len(col_headers) + "\\\\")
            continue

        vals = []
        am = row_data.get('all_mean', np.nan)
        asd = row_data.get('all_sd', np.nan)
        tm = row_data.get('t_mean', np.nan)
        tsd = row_data.get('t_sd', np.nan)
        cm = row_data.get('c_mean', np.nan)
        csd = row_data.get('c_sd', np.nan)
        d = row_data.get('diff', np.nan)
        se = row_data.get('se', np.nan)
        stars = row_data.get('stars', '')

        vals = [fmt(am), f"[{fmt(asd)}]" if not pd.isna(asd) else '',
                fmt(tm), f"[{fmt(tsd)}]" if not pd.isna(tsd) else '',
                fmt(cm), f"[{fmt(csd)}]" if not pd.isna(csd) else '',
                f"{fmt(d)}{stars}" if not pd.isna(d) else '',
                f"({fmt(se)})" if not pd.isna(se) else '']

        lines.append(f"\\quad {row_data['label']} & " + " & ".join(vals) + " \\\\")

    lines.append("\\bottomrule")
    lines.append("\\end{tabular}")
    lines.append("\\begin{tablenotes}")
    lines.append("\\footnotesize")
    lines.append(f"\\item {note}")
    lines.append("\\end{tablenotes}")
    lines.append("\\end{table}")
    return "\n".join(lines)

# Table 1 TeX
tex1 = write_tex_summary_table(
    table1_rows,
    "Summary Statistics of Individuals (Age 0--40 at Conflict Start)",
    "tab:summary_stats",
    ['All', '[SD]', 'Treatment', '[SD]', 'Control', '[SD]', 'Diff', '(SE)'],
    "Treatment = aged 0--17 at conflict start. Control = aged 18--40. $^{***}$ p$<$0.01, $^{**}$ p$<$0.05, $^{*}$ p$<$0.10"
)

# Table 3 TeX (in utero males)
tex3 = write_tex_summary_table(
    table3_rows,
    "Summary Statistics: Male Children In Utero During Conflict (Born 1996--2006)",
    "tab:males_in_utero",
    ['All', '[SD]', 'High Confl.', '[SD]', 'Low Confl.', '[SD]', 'Diff', '(SE)'],
    "Males born 1996--2006. Treatment = above-median casualties. $^{***}$ p$<$0.01, $^{**}$ p$<$0.05, $^{*}$ p$<$0.10"
)

# Table 4 TeX (DID regression - Phadera cohorts)
tex4_lines = []
tex4_lines.append("\\begin{table}[htbp]")
tex4_lines.append("\\centering")
tex4_lines.append("\\footnotesize")
tex4_lines.append("\\caption{Impact on Migration by Age at Start of the Civil War (Equation 7)}")
tex4_lines.append("\\label{tab:did_phadera}")
tex4_lines.append("\\begin{tabular}{l cccc}")
tex4_lines.append("\\toprule")
tex4_lines.append(" & \\multicolumn{2}{c}{International Migration} & \\multicolumn{2}{c}{Domestic Migration} \\\\")
tex4_lines.append("\\cmidrule(lr){2-3} \\cmidrule(lr){4-5}")
tex4_lines.append(" & (1) Months & (2) Casualties & (3) Months & (4) Casualties \\\\")
tex4_lines.append("\\midrule")

for cohort_suffix, row_label in cohort_interactions:
    coef_parts = [f"\\quad {row_label}"]
    se_parts = [""]
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_phadera'
            inter_name = f'{conf_var}_x_{cohort_suffix}'
            if key in results_table4 and inter_name in results_table4[key]:
                res = results_table4[key][inter_name]
                coef_parts.append(f"{res['coef']:.6f}$^{{{res['stars']}}}$" if res['stars'] else f"{res['coef']:.6f}")
                se_parts.append(f"({res['se']:.6f})")
            else:
                coef_parts.append("")
                se_parts.append("")
    tex4_lines.append(" & ".join(coef_parts) + " \\\\")
    tex4_lines.append(" & ".join(se_parts) + " \\\\")

tex4_lines.append("\\midrule")
# Bottom stats
for stat_label, stat_key, fmt_str in [('Observations', 'nobs', '{:.0f}'), ('Adj. R-squared', 'r2_adj', '{:.4f}'), ('Control Mean (Age 16--21)', 'ctrl_mean', '{:.4f}')]:
    parts = [stat_label]
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_phadera'
            if key in results_table4:
                val = results_table4[key].get(stat_key, np.nan)
                parts.append(fmt_str.format(val) if not np.isnan(val) else '')
            else:
                parts.append('')
    tex4_lines.append(" & ".join(parts) + " \\\\")

tex4_lines.append("District FE & Yes & Yes & Yes & Yes \\\\")
tex4_lines.append("Cohort dummies & Yes & Yes & Yes & Yes \\\\")
tex4_lines.append("\\bottomrule")
tex4_lines.append("\\end{tabular}")
tex4_lines.append("\\begin{tablenotes}")
tex4_lines.append("\\footnotesize")
tex4_lines.append("\\item Standard errors clustered at PSU level in parentheses. $^{***}$ p$<$0.01, $^{**}$ p$<$0.05, $^{*}$ p$<$0.10")
tex4_lines.append("\\end{tablenotes}")
tex4_lines.append("\\end{table}")
tex4 = "\n".join(tex4_lines)

# Table 5 TeX (DID with our age groups)
tex5_lines = []
tex5_lines.append("\\begin{table}[htbp]")
tex5_lines.append("\\centering")
tex5_lines.append("\\footnotesize")
tex5_lines.append("\\caption{Impact on Migration by Age Group at Start of the Civil War}")
tex5_lines.append("\\label{tab:did_agegroups}")
tex5_lines.append("\\begin{tabular}{l cccc}")
tex5_lines.append("\\toprule")
tex5_lines.append(" & \\multicolumn{2}{c}{International Migration} & \\multicolumn{2}{c}{Domestic Migration} \\\\")
tex5_lines.append("\\cmidrule(lr){2-3} \\cmidrule(lr){4-5}")
tex5_lines.append(" & (1) Months & (2) Casualties & (3) Months & (4) Casualties \\\\")
tex5_lines.append("\\midrule")

for cohort_suffix, row_label in cohort_interactions2:
    coef_parts = [f"\\quad {row_label}"]
    se_parts = [""]
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_agegroups'
            inter_name = f'{conf_var}_x_{cohort_suffix}'
            if key in results_table4 and inter_name in results_table4[key]:
                res = results_table4[key][inter_name]
                coef_parts.append(f"{res['coef']:.6f}$^{{{res['stars']}}}$" if res['stars'] else f"{res['coef']:.6f}")
                se_parts.append(f"({res['se']:.6f})")
            else:
                coef_parts.append("")
                se_parts.append("")
    tex5_lines.append(" & ".join(coef_parts) + " \\\\")
    tex5_lines.append(" & ".join(se_parts) + " \\\\")

tex5_lines.append("\\midrule")
for stat_label, stat_key, fmt_str in [('Observations', 'nobs', '{:.0f}'), ('Adj. R-squared', 'r2_adj', '{:.4f}'), ('Control Mean (Age 19--25)', 'ctrl_mean', '{:.4f}')]:
    parts = [stat_label]
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            key = f'{outcome}_{conf_var}_agegroups'
            if key in results_table4:
                val = results_table4[key].get(stat_key, np.nan)
                parts.append(fmt_str.format(val) if not np.isnan(val) else '')
            else:
                parts.append('')
    tex5_lines.append(" & ".join(parts) + " \\\\")

tex5_lines.append("District FE & Yes & Yes & Yes & Yes \\\\")
tex5_lines.append("Cohort dummies & Yes & Yes & Yes & Yes \\\\")
tex5_lines.append("\\bottomrule")
tex5_lines.append("\\end{tabular}")
tex5_lines.append("\\begin{tablenotes}")
tex5_lines.append("\\footnotesize")
tex5_lines.append("\\item Control group: Age 19--25 at conflict start. Standard errors clustered at PSU level. $^{***}$ p$<$0.01, $^{**}$ p$<$0.05, $^{*}$ p$<$0.10")
tex5_lines.append("\\end{tablenotes}")
tex5_lines.append("\\end{table}")
tex5 = "\n".join(tex5_lines)

# Write all TeX
full_tex = tex1 + "\n\n\\clearpage\n\n" + tex3 + "\n\n\\clearpage\n\n" + tex4 + "\n\n\\clearpage\n\n" + tex5
tex_path = f"{OUT_DIR}/phadera_replication_tables.tex"
with open(tex_path, 'w') as f:
    f.write(full_tex)
print(f"  TEX saved: {tex_path}")

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE PDF
# ══════════════════════════════════════════════════════════════════════════════
print("  GENERATING PDF output...")

from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas as pdf_canvas

NAVY_C = rl_colors.HexColor("#1B2A4A")
TEAL_C = rl_colors.HexColor("#2E8B8B")
NOTE_C = rl_colors.HexColor("#888888")
LGRAY_C = rl_colors.HexColor("#F5F5F5")
GLINE_C = rl_colors.HexColor("#D0D0D0")

def draw_summary_pdf(cv, title, rows, col_headers, pw, ph, note_text):
    """Draw a summary statistics table on a PDF page."""
    left = 30; right = pw - 30; tw = right - left; top = ph - 30
    y = top

    cv.setFont("Helvetica-Bold", 11)
    cv.setFillColor(NAVY_C)
    cv.drawString(left, y, title)
    y -= 20

    # Column headers
    n_cols = len(col_headers)
    label_w = 0.22 * tw
    data_w = (tw - label_w) / n_cols

    cv.setFont("Helvetica-Bold", 7)
    cv.setFillColor(NAVY_C)
    for ci, hdr in enumerate(col_headers):
        x = left + label_w + ci * data_w
        cv.drawCentredString(x + data_w/2, y, hdr)
    y -= 3
    cv.setStrokeColor(NAVY_C); cv.setLineWidth(1)
    cv.line(left, y, right, y)
    y -= 12

    ROW_H = 13
    alt = False
    for row_data in rows:
        if row_data.get('section'):
            cv.setFillColor(rl_colors.HexColor("#E8F5F3"))
            cv.rect(left, y - ROW_H + 4, tw, ROW_H, fill=1, stroke=0)
            cv.setFont("Helvetica-BoldOblique", 8)
            cv.setFillColor(TEAL_C)
            cv.drawString(left + 4, y - 4, row_data['label'])
            y -= ROW_H
            alt = False
            continue

        if alt:
            cv.setFillColor(LGRAY_C)
            cv.rect(left, y - ROW_H + 4, tw, ROW_H, fill=1, stroke=0)

        cv.setStrokeColor(GLINE_C); cv.setLineWidth(0.3)
        cv.line(left, y - ROW_H + 4, right, y - ROW_H + 4)

        cv.setFont("Helvetica", 7.5)
        cv.setFillColor(rl_colors.HexColor("#1A1A1A"))
        cv.drawString(left + 8, y - 4, row_data['label'])

        vals = [fmt(row_data.get('all_mean')), f"[{fmt(row_data.get('all_sd'))}]" if not pd.isna(row_data.get('all_sd', np.nan)) else '',
                fmt(row_data.get('t_mean')), f"[{fmt(row_data.get('t_sd'))}]" if not pd.isna(row_data.get('t_sd', np.nan)) else '',
                fmt(row_data.get('c_mean')), f"[{fmt(row_data.get('c_sd'))}]" if not pd.isna(row_data.get('c_sd', np.nan)) else '',
                f"{fmt(row_data.get('diff'))}{row_data.get('stars','')}" if not pd.isna(row_data.get('diff', np.nan)) else '',
                f"({fmt(row_data.get('se'))})" if not pd.isna(row_data.get('se', np.nan)) else '']

        for ci, v in enumerate(vals):
            if v:
                x = left + label_w + ci * data_w
                cv.drawCentredString(x + data_w/2, y - 4, str(v))

        y -= ROW_H
        alt = not alt

    cv.setStrokeColor(NAVY_C); cv.setLineWidth(1)
    cv.line(left, y + 4, right, y + 4)

    y -= 14
    cv.setFont("Helvetica", 6)
    cv.setFillColor(NOTE_C)
    cv.drawString(left, y, note_text)

# Create PDF
pdf_path = f"{OUT_DIR}/phadera_replication_tables.pdf"
page_size = landscape(letter)
pw, ph = page_size
c = pdf_canvas.Canvas(pdf_path, pagesize=page_size)

# Page 1: Table 1
draw_summary_pdf(c, 'Table 1: Summary Statistics (Age 0-40 at Conflict Start)',
                  table1_rows, ['All Mean', '[SD]', 'Treat Mean', '[SD]', 'Ctrl Mean', '[SD]', 'Diff', '(SE)'],
                  pw, ph, "Treatment=Age 0-17 at conflict start. Control=Age 18-40. *** p<0.01, ** p<0.05, * p<0.10")
c.showPage()

# Page 2: Table 3 (males in utero)
draw_summary_pdf(c, 'Table 3: Male Children In Utero During Conflict (Born 1996-2006)',
                  table3_rows, ['All Mean', '[SD]', 'High Confl.', '[SD]', 'Low Confl.', '[SD]', 'Diff', '(SE)'],
                  pw, ph, "Males born 1996-2006. Treatment=above-median casualties. *** p<0.01, ** p<0.05, * p<0.10")

c.save()
print(f"  PDF saved: {pdf_path}")

# Copy figures to results too
import shutil
for fig_name in ['figure_4_cohort_diagram.png', 'figure_4_cohort_diagram.pdf',
                  'figure_5_migration_by_conflict.png', 'figure_5_migration_by_conflict.pdf']:
    src = f"{FIG_DIR}/{fig_name}"
    dst = f"{OUT_DIR}/{fig_name}"
    if os.path.exists(src):
        shutil.copy2(src, dst)

print("\n" + "="*70)
print("  ALL DONE!")
print("="*70)
print(f"  Results: {OUT_DIR}/")
print(f"  Figures: {FIG_DIR}/")
