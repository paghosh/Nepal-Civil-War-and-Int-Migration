"""
eq8_intergenerational.py

Replicates Phadera (2021) Equation 8 - Intergenerational specification.
Child outcome = f(mother's conflict exposure × mother's cohort) + child controls

Adaptation for NLFS data with migration as outcome:
- Link mothers (female head/spouse/parent) to children (son/daughter/grandchild) within households
- Mother's conflict exposure = months of war / casualties in district
- Mother's cohort = age at conflict start (same as Eq 7)
- Child outcomes = international migration, domestic migration
- Child controls = child birth year FE, child sex (girl indicator), birth order FE
"""

import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.formula.api as smf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors as rl_colors
from reportlab.pdfgen import canvas as pdf_canvas
import warnings, os
warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════════════════════
BASE = "/Users/pallab.ghosh/Documents/GitHub/Nepal-Civil-War-and-Int-Migration"
DATA_PATH = f"{BASE}/data/Modified_Data/1_conflict_present_absentee_data.dta"
OUT_DIR = f"{BASE}/tables/pallab/march_2026/summary/results"

SURVEY_YEAR = 2017
CONFLICT_START = 1996
CONFLICT_END = 2006

# ══════════════════════════════════════════════════════════════════════════════
# LOAD & PREPARE DATA
# ══════════════════════════════════════════════════════════════════════════════
print("Loading data...")
df = pd.read_stata(DATA_PATH, convert_categoricals=False)

df['sex_num'] = pd.to_numeric(df['sex'], errors='coerce')
df['male'] = (df['sex_num'] == 1).astype(float)
df['female'] = (df['sex_num'] == 2).astype(float)
df['girl'] = df['female']  # child-level control
df['age_num'] = pd.to_numeric(df['age'], errors='coerce')
df['birth_year'] = SURVEY_YEAR - df['age_num']
df['age_at_conflict_start'] = CONFLICT_START - df['birth_year']

# Casualty variables
for cv in ['cas_own_any','cas_own_fatal','mwar_own_any','mwar_own_fatal',
           'cas_nbr_any','cas_nbr_fatal','mwar_nbr_any','mwar_nbr_fatal']:
    df[cv] = pd.to_numeric(df[cv], errors='coerce')

dist_cas = df.dropna(subset=['cas_own_any']).groupby('dist')[
    ['cas_own_any','cas_own_fatal','mwar_own_any','mwar_own_fatal',
     'cas_nbr_any','cas_nbr_fatal','mwar_nbr_any','mwar_nbr_fatal']].first()
for cv in dist_cas.columns:
    missing = df[cv].isna()
    df.loc[missing, cv] = df.loc[missing, 'dist'].map(dist_cas[cv])

# Ethnicity
caste = pd.to_numeric(df['caste'], errors='coerce')
df['caste_num'] = caste
df['caste_num'] = df.groupby(['psu','hhld'])['caste_num'].transform(lambda x: x.ffill().bfill())
hill_high={1,2,14,20,27,48,49}
def get_high(c):
    if pd.isna(c): return 0.0
    return 1.0 if int(c) in hill_high else 0.0
df['high_caste'] = df['caste_num'].apply(get_high)

# Migration outcomes
df['intl_mig'] = df['international_absentee_only'].copy()
df['dom_mig'] = df['national'].copy()

# District/PSU numeric
df['dist_num'] = pd.to_numeric(df['dist'], errors='coerce')
df['psu_num'] = pd.to_numeric(df['psu'], errors='coerce')
df['rel_hhh'] = pd.to_numeric(df['rel_hhh'], errors='coerce')

print(f"Data loaded: {len(df)} rows")

# ══════════════════════════════════════════════════════════════════════════════
# LINK MOTHERS TO CHILDREN
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  LINKING MOTHERS TO CHILDREN WITHIN HOUSEHOLDS")
print("="*70)

# Mothers: female (sex=2), relationship = head(1), spouse(2), or parent(6)
# Children: relationship = son/daughter(3), daughter/son-in-law(4), grandchild(5)
# We focus on children who might be migrants (any age)

mothers = df[(df['sex_num'] == 2) & (df['rel_hhh'].isin([1, 2, 6]))].copy()
mothers = mothers.rename(columns={
    'age_num': 'mother_age', 'birth_year': 'mother_birth_year',
    'age_at_conflict_start': 'mother_age_at_cs',
    'high_caste': 'mother_high_caste',
})
mothers['mother_idx'] = mothers.index
mother_cols = ['psu', 'hhld', 'mother_idx', 'mother_age', 'mother_birth_year',
               'mother_age_at_cs', 'mother_high_caste', 'dist_num', 'psu_num',
               'mwar_own_any', 'cas_own_any', 'mwar_own_fatal', 'cas_own_fatal']
mothers_slim = mothers[mother_cols].copy()

# Children: son/daughter(3), son/daughter-in-law(4), grandchild(5) - all genders
children = df[df['rel_hhh'].isin([3, 4, 5])].copy()
children['child_idx'] = children.index
child_cols = ['psu', 'hhld', 'child_idx', 'age_num', 'birth_year', 'sex_num',
              'girl', 'intl_mig', 'dom_mig', 'male']
children_slim = children[child_cols].copy()
children_slim = children_slim.rename(columns={
    'age_num': 'child_age', 'birth_year': 'child_birth_year',
    'sex_num': 'child_sex'
})

# Merge: match children to mothers within same household
# If multiple mothers in household, take the one most likely to be the biological mother
# (spouse or head, and age gap 15-50 years)
mc = children_slim.merge(mothers_slim, on=['psu', 'hhld'], how='inner')
mc['age_gap'] = mc['mother_age'] - mc['child_age']
mc = mc[(mc['age_gap'] >= 15) & (mc['age_gap'] <= 50)]  # plausible mother-child age gap

# If multiple mothers match, keep the one with smallest age gap (most likely biological)
mc['age_gap_abs'] = (mc['age_gap'] - 25).abs()  # typical motherhood age ~25
mc = mc.sort_values('age_gap_abs').drop_duplicates(subset='child_idx', keep='first')

# Birth order within household (by child age, oldest = 1)
mc = mc.sort_values(['psu', 'hhld', 'mother_idx', 'child_age'], ascending=[True, True, True, False])
mc['birth_order'] = mc.groupby(['psu', 'hhld', 'mother_idx']).cumcount() + 1
mc['birth_order'] = mc['birth_order'].clip(upper=6)  # cap at 6

print(f"Mother-child pairs: {len(mc)}")
print(f"Unique mothers: {mc['mother_idx'].nunique()}")
print(f"Unique children: {mc['child_idx'].nunique()}")

# ══════════════════════════════════════════════════════════════════════════════
# CREATE MOTHER'S COHORT VARIABLES
# ══════════════════════════════════════════════════════════════════════════════

# Mother's cohort based on her age at conflict start (following Phadera)
mc['m_cohort_0_3'] = mc['mother_age_at_cs'].between(0, 3).astype(float)
mc['m_cohort_4_8'] = mc['mother_age_at_cs'].between(4, 8).astype(float)
mc['m_cohort_9_15'] = mc['mother_age_at_cs'].between(9, 15).astype(float)
mc['m_cohort_16_21'] = mc['mother_age_at_cs'].between(16, 21).astype(float)  # control
mc['m_cohort_22_29'] = mc['mother_age_at_cs'].between(22, 29).astype(float)  # placebo

# Our age groups for mother's cohort
mc['m_cohort_0_5'] = mc['mother_age_at_cs'].between(0, 5).astype(float)
mc['m_cohort_6_10'] = mc['mother_age_at_cs'].between(6, 10).astype(float)
mc['m_cohort_11_18'] = mc['mother_age_at_cs'].between(11, 18).astype(float)
mc['m_cohort_19_25'] = mc['mother_age_at_cs'].between(19, 25).astype(float)  # control
mc['m_cohort_26_40'] = mc['mother_age_at_cs'].between(26, 40).astype(float)  # placebo

# Interactions: mother's conflict × mother's cohort
for conflict_var in ['mwar_own_any', 'cas_own_any']:
    for cohort_name in ['0_3', '4_8', '9_15', '22_29',
                         '0_5', '6_10', '11_18', '26_40']:
        mc[f'{conflict_var}_x_m{cohort_name}'] = (
            mc[conflict_var] * mc[f'm_cohort_{cohort_name}']
        )

# Filter to mothers aged 0-40 at conflict start (to have meaningful cohort assignment)
mc_sample = mc[(mc['mother_age_at_cs'].between(0, 40)) &
               mc['mwar_own_any'].notna() &
               mc['child_age'].notna()].copy()

print(f"Regression sample (mothers age 0-40 at CS): {len(mc_sample)}")
print(f"  Mother age at conflict start: {mc_sample['mother_age_at_cs'].describe()}")
print(f"  Child age: {mc_sample['child_age'].describe()}")

# ══════════════════════════════════════════════════════════════════════════════
# RUN EQUATION 8 REGRESSIONS
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  RUNNING EQUATION 8: INTERGENERATIONAL REGRESSIONS")
print("="*70)

results_eq8 = {}

for outcome in ['intl_mig', 'dom_mig']:
    for conflict_var in ['mwar_own_any', 'cas_own_any']:
        # --- Phadera cohorts ---
        interactions = [f'{conflict_var}_x_m22_29', f'{conflict_var}_x_m9_15',
                        f'{conflict_var}_x_m4_8', f'{conflict_var}_x_m0_3']

        # Eq 8: child outcome ~ mother's conflict×cohort + mother's cohort FE
        #   + mother's high caste + child girl + child birth order FE + district FE
        formula = f'{outcome} ~ ' + ' + '.join(interactions)
        formula += ' + m_cohort_0_3 + m_cohort_4_8 + m_cohort_9_15 + m_cohort_22_29'
        formula += ' + mother_high_caste + girl + C(birth_order) + C(dist_num)'

        try:
            valid = mc_sample.dropna(subset=[outcome] + interactions +
                                     ['mother_high_caste', 'girl', 'birth_order', 'dist_num'])
            if len(valid) < 50:
                print(f"  SKIP {outcome} ~ {conflict_var} (Phadera): N={len(valid)} too small")
                continue

            model = smf.ols(formula, data=valid).fit(
                cov_type='cluster', cov_kwds={'groups': valid['psu_num']})

            key = f'{outcome}_{conflict_var}_phadera_eq8'
            results_eq8[key] = {}
            for interaction in interactions:
                coef = model.params.get(interaction, np.nan)
                se = model.bse.get(interaction, np.nan)
                pval = model.pvalues.get(interaction, np.nan)
                stars = '***' if pval < 0.01 else ('**' if pval < 0.05 else ('*' if pval < 0.10 else ''))
                results_eq8[key][interaction] = {
                    'coef': coef, 'se': se, 'pval': pval, 'stars': stars
                }
            results_eq8[key]['nobs'] = model.nobs
            results_eq8[key]['r2_adj'] = model.rsquared_adj

            ctrl_mask = valid['m_cohort_16_21'] == 1
            ctrl_mean = valid.loc[ctrl_mask, outcome].mean() if ctrl_mask.any() else np.nan
            results_eq8[key]['ctrl_mean'] = ctrl_mean

            print(f"  {outcome} ~ {conflict_var} (Phadera Eq8): N={model.nobs:.0f}, R2={model.rsquared_adj:.4f}")
        except Exception as e:
            print(f"  ERROR: {outcome} ~ {conflict_var} Eq8 Phadera: {e}")

        # --- Our age groups ---
        interactions2 = [f'{conflict_var}_x_m26_40', f'{conflict_var}_x_m11_18',
                         f'{conflict_var}_x_m6_10', f'{conflict_var}_x_m0_5']

        formula2 = f'{outcome} ~ ' + ' + '.join(interactions2)
        formula2 += ' + m_cohort_0_5 + m_cohort_6_10 + m_cohort_11_18 + m_cohort_26_40'
        formula2 += ' + mother_high_caste + girl + C(birth_order) + C(dist_num)'

        try:
            valid2 = mc_sample.dropna(subset=[outcome] + interactions2 +
                                       ['mother_high_caste', 'girl', 'birth_order', 'dist_num'])
            if len(valid2) < 50:
                print(f"  SKIP {outcome} ~ {conflict_var} (age groups Eq8): N={len(valid2)} too small")
                continue

            model2 = smf.ols(formula2, data=valid2).fit(
                cov_type='cluster', cov_kwds={'groups': valid2['psu_num']})

            key2 = f'{outcome}_{conflict_var}_agegroups_eq8'
            results_eq8[key2] = {}
            for interaction in interactions2:
                coef = model2.params.get(interaction, np.nan)
                se = model2.bse.get(interaction, np.nan)
                pval = model2.pvalues.get(interaction, np.nan)
                stars = '***' if pval < 0.01 else ('**' if pval < 0.05 else ('*' if pval < 0.10 else ''))
                results_eq8[key2][interaction] = {
                    'coef': coef, 'se': se, 'pval': pval, 'stars': stars
                }
            results_eq8[key2]['nobs'] = model2.nobs
            results_eq8[key2]['r2_adj'] = model2.rsquared_adj

            ctrl_mask2 = valid2['m_cohort_19_25'] == 1
            ctrl_mean2 = valid2.loc[ctrl_mask2, outcome].mean() if ctrl_mask2.any() else np.nan
            results_eq8[key2]['ctrl_mean'] = ctrl_mean2

            print(f"  {outcome} ~ {conflict_var} (Age Groups Eq8): N={model2.nobs:.0f}, R2={model2.rsquared_adj:.4f}")
        except Exception as e:
            print(f"  ERROR: {outcome} ~ {conflict_var} Eq8 age groups: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# WRITE RESULTS TO EXCEL (append to existing workbook)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*70)
print("  WRITING EQUATION 8 RESULTS")
print("="*70)

from openpyxl import load_workbook

# Style constants
NAVY = "1B2A4A"
WHITE = "FFFFFF"
LGRAY = "F5F5F5"
MGRAY = "E0E0E0"

hdr_font = Font(name='Arial', size=10, bold=True, color=WHITE)
hdr_fill = PatternFill('solid', fgColor=NAVY)
data_font = Font(name='Arial', size=9, color='1A1A1A')
note_font = Font(name='Arial', size=7.5, italic=True, color='666666')

outcome_labels = {'intl_mig': 'International Migration', 'dom_mig': 'Domestic Migration'}
conflict_labels = {'mwar_own_any': 'Months of War', 'cas_own_any': 'Casualty Count'}

def fmt(val, decimals=2):
    if pd.isna(val) or val is None: return ''
    if isinstance(val, int) or (isinstance(val, float) and val == int(val) and abs(val) > 100):
        return f'{int(val):,}'
    return f'{val:.{decimals}f}'

def write_regression_sheet(wb, sheet_name, title, cohort_interactions, results_dict, suffix, ctrl_label):
    """Write a regression results sheet to the workbook."""
    ws = wb.create_sheet(sheet_name)
    r = 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(row=r, column=1, value=title).font = Font(name='Arial', size=12, bold=True, color=NAVY)
    r += 2

    # Column headers
    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        c = ws.cell(row=r, column=col, value=outcome_labels[outcome])
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=col+1).fill = hdr_fill
        col += 2
    r += 1

    col = 2
    for outcome in ['intl_mig', 'dom_mig']:
        for conf_var in ['mwar_own_any', 'cas_own_any']:
            c = ws.cell(row=r, column=col, value=f'({col-1})\n{conflict_labels[conf_var]}')
            c.font = Font(name='Arial', size=8, bold=True, color=NAVY)
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = Border(bottom=Side(style='medium', color=NAVY))
            col += 1
    r += 1

    # Regression rows
    for cohort_suffix, row_label in cohort_interactions:
        # Coefficient row
        ws.cell(row=r, column=1, value=row_label).font = Font(name='Arial', size=9, bold=True)
        col = 2
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                inter_name = f'{conf_var}_x_m{cohort_suffix}'
                if key in results_dict and inter_name in results_dict[key]:
                    res = results_dict[key][inter_name]
                    val_str = f"{res['coef']:.6f}{res['stars']}"
                    ws.cell(row=r, column=col, value=val_str).font = data_font
                    ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
                col += 1
        r += 1

        # SE row
        col = 2
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                inter_name = f'{conf_var}_x_m{cohort_suffix}'
                if key in results_dict and inter_name in results_dict[key]:
                    res = results_dict[key][inter_name]
                    ws.cell(row=r, column=col, value=f"({res['se']:.6f})").font = Font(name='Arial', size=8, color='666666')
                    ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
                col += 1
        r += 1

    # Bottom stats
    r += 1
    for stat_label, stat_key in [('Observations', 'nobs'), ('Adjusted R-squared', 'r2_adj'),
                                   (f'Control Mean ({ctrl_label})', 'ctrl_mean')]:
        ws.cell(row=r, column=1, value=stat_label).font = Font(name='Arial', size=9, bold=True)
        col = 2
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                if key in results_dict:
                    val = results_dict[key].get(stat_key, np.nan)
                    if stat_key == 'nobs':
                        ws.cell(row=r, column=col, value=f"{int(val):,}" if not np.isnan(val) else '').font = data_font
                    else:
                        ws.cell(row=r, column=col, value=f"{val:.4f}" if not np.isnan(val) else '').font = data_font
                    ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
                col += 1
        r += 1

    # Notes
    r += 1
    ws.cell(row=r, column=1, value="Controls: District FE, mother's cohort dummies, mother's high caste, child girl indicator, birth order FE").font = note_font
    ws.cell(row=r+1, column=1, value="Standard errors clustered at PSU level. *** p<0.01, ** p<0.05, * p<0.10").font = note_font
    ws.cell(row=r+2, column=1, value="Intergenerational specification (Equation 8): child migration ~ mother's conflict exposure x mother's cohort").font = note_font

    ws.column_dimensions['A'].width = 38
    for ci in range(2, 7):
        ws.column_dimensions[chr(64+ci)].width = 18

# Load existing workbook and add Eq 8 sheets
xlsx_path = f"{OUT_DIR}/phadera_replication_tables.xlsx"
wb = load_workbook(xlsx_path)

# Table 6: Eq 8 with Phadera cohorts
phadera_cohort_interactions = [
    ('22_29', "Mother Age 22-29 x Conflict (Placebo)"),
    ('9_15', "Mother Age 9-15 x Conflict"),
    ('4_8', "Mother Age 4-8 x Conflict"),
    ('0_3', "Mother Age 0-3 x Conflict"),
]
write_regression_sheet(wb, "Table 6 - Eq8 Phadera",
    "Table 6: Intergenerational Impact on Children's Migration (Equation 8, Phadera Cohorts)",
    phadera_cohort_interactions, results_eq8, 'phadera_eq8', "Mother Age 16-21")

# Table 7: Eq 8 with our age groups
our_cohort_interactions = [
    ('26_40', "Mother Age 26-40 x Conflict (Placebo)"),
    ('11_18', "Mother Age 11-18 x Conflict"),
    ('6_10', "Mother Age 6-10 x Conflict"),
    ('0_5', "Mother Age 0-5 x Conflict"),
]
write_regression_sheet(wb, "Table 7 - Eq8 Age Groups",
    "Table 7: Intergenerational Impact on Children's Migration (Equation 8, Our Age Groups)",
    our_cohort_interactions, results_eq8, 'agegroups_eq8', "Mother Age 19-25")

wb.save(xlsx_path)
print(f"  Updated XLSX: {xlsx_path}")

# ══════════════════════════════════════════════════════════════════════════════
# APPEND TO LaTeX FILE
# ══════════════════════════════════════════════════════════════════════════════
print("  Appending Eq 8 to LaTeX...")

def write_tex_regression(cohort_interactions, results_dict, suffix, title, label, ctrl_label):
    lines = []
    lines.append("\\begin{table}[htbp]")
    lines.append("\\centering")
    lines.append("\\footnotesize")
    lines.append(f"\\caption{{{title}}}")
    lines.append(f"\\label{{{label}}}")
    lines.append("\\begin{tabular}{l cccc}")
    lines.append("\\toprule")
    lines.append(" & \\multicolumn{2}{c}{International Migration} & \\multicolumn{2}{c}{Domestic Migration} \\\\")
    lines.append("\\cmidrule(lr){2-3} \\cmidrule(lr){4-5}")
    lines.append(" & (1) Months & (2) Casualties & (3) Months & (4) Casualties \\\\")
    lines.append("\\midrule")

    for cohort_suffix, row_label in cohort_interactions:
        coef_parts = [f"\\quad {row_label}"]
        se_parts = [""]
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                inter_name = f'{conf_var}_x_m{cohort_suffix}'
                if key in results_dict and inter_name in results_dict[key]:
                    res = results_dict[key][inter_name]
                    coef_parts.append(f"{res['coef']:.6f}$^{{{res['stars']}}}$" if res['stars'] else f"{res['coef']:.6f}")
                    se_parts.append(f"({res['se']:.6f})")
                else:
                    coef_parts.append("")
                    se_parts.append("")
        lines.append(" & ".join(coef_parts) + " \\\\")
        lines.append(" & ".join(se_parts) + " \\\\")

    lines.append("\\midrule")
    for stat_label, stat_key, fmt_str in [('Observations', 'nobs', '{:.0f}'),
                                            ('Adj. R-squared', 'r2_adj', '{:.4f}'),
                                            (f'Control Mean ({ctrl_label})', 'ctrl_mean', '{:.4f}')]:
        parts = [stat_label]
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                if key in results_dict:
                    val = results_dict[key].get(stat_key, np.nan)
                    parts.append(fmt_str.format(val) if not np.isnan(val) else '')
                else:
                    parts.append('')
        lines.append(" & ".join(parts) + " \\\\")

    lines.append("District FE & Yes & Yes & Yes & Yes \\\\")
    lines.append("Mother's cohort FE & Yes & Yes & Yes & Yes \\\\")
    lines.append("Child controls & Yes & Yes & Yes & Yes \\\\")
    lines.append("\\bottomrule")
    lines.append("\\end{tabular}")
    lines.append("\\begin{tablenotes}")
    lines.append("\\footnotesize")
    lines.append("\\item Intergenerational specification (Equation 8). Child controls: girl indicator, birth order FE.")
    lines.append("\\item Standard errors clustered at PSU level. $^{***}$ p$<$0.01, $^{**}$ p$<$0.05, $^{*}$ p$<$0.10")
    lines.append("\\end{tablenotes}")
    lines.append("\\end{table}")
    return "\n".join(lines)

tex6 = write_tex_regression(
    phadera_cohort_interactions, results_eq8, 'phadera_eq8',
    "Intergenerational Impact on Children's Migration (Equation 8, Phadera Cohorts)",
    "tab:eq8_phadera", "Mother Age 16--21")

tex7 = write_tex_regression(
    our_cohort_interactions, results_eq8, 'agegroups_eq8',
    "Intergenerational Impact on Children's Migration (Equation 8, Our Age Groups)",
    "tab:eq8_agegroups", "Mother Age 19--25")

# Append to existing tex file
tex_path = f"{OUT_DIR}/phadera_replication_tables.tex"
with open(tex_path, 'a') as f:
    f.write("\n\n\\clearpage\n\n" + tex6 + "\n\n\\clearpage\n\n" + tex7)
print(f"  Updated TEX: {tex_path}")

# ══════════════════════════════════════════════════════════════════════════════
# APPEND TO PDF
# ══════════════════════════════════════════════════════════════════════════════
print("  Generating Eq 8 PDF pages...")

NAVY_C = rl_colors.HexColor("#1B2A4A")
TEAL_C = rl_colors.HexColor("#2E8B8B")
NOTE_C = rl_colors.HexColor("#888888")
LGRAY_C = rl_colors.HexColor("#F5F5F5")
GLINE_C = rl_colors.HexColor("#D0D0D0")

def draw_regression_pdf(cv, title, cohort_interactions, results_dict, suffix, ctrl_label, pw, ph):
    """Draw regression results on a PDF page."""
    left = 30; right = pw - 30; tw = right - left; top = ph - 30
    y = top

    cv.setFont("Helvetica-Bold", 11)
    cv.setFillColor(NAVY_C)
    cv.drawString(left, y, title)
    y -= 18

    # Column headers
    cols = ['Intl Mig\n(Months)', 'Intl Mig\n(Casualties)', 'Dom Mig\n(Months)', 'Dom Mig\n(Casualties)']
    label_w = 0.35 * tw
    data_w = (tw - label_w) / len(cols)

    cv.setFont("Helvetica-Bold", 8)
    cv.setFillColor(NAVY_C)
    for ci, hdr in enumerate(cols):
        x = left + label_w + ci * data_w
        cv.drawCentredString(x + data_w/2, y, hdr)
    y -= 3
    cv.setStrokeColor(NAVY_C); cv.setLineWidth(1)
    cv.line(left, y, right, y)
    y -= 14

    ROW_H = 13
    for cohort_suffix, row_label in cohort_interactions:
        # Coef row
        cv.setFont("Helvetica-Bold", 7.5)
        cv.setFillColor(rl_colors.HexColor("#1A1A1A"))
        cv.drawString(left + 4, y - 4, row_label)

        cv.setFont("Helvetica", 7.5)
        col_idx = 0
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                inter_name = f'{conf_var}_x_m{cohort_suffix}'
                if key in results_dict and inter_name in results_dict[key]:
                    res = results_dict[key][inter_name]
                    x = left + label_w + col_idx * data_w
                    cv.drawCentredString(x + data_w/2, y - 4, f"{res['coef']:.6f}{res['stars']}")
                col_idx += 1
        y -= ROW_H

        # SE row
        cv.setFont("Helvetica", 6.5)
        cv.setFillColor(NOTE_C)
        col_idx = 0
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                inter_name = f'{conf_var}_x_m{cohort_suffix}'
                if key in results_dict and inter_name in results_dict[key]:
                    res = results_dict[key][inter_name]
                    x = left + label_w + col_idx * data_w
                    cv.drawCentredString(x + data_w/2, y - 4, f"({res['se']:.6f})")
                col_idx += 1
        y -= ROW_H

    cv.setStrokeColor(NAVY_C); cv.setLineWidth(0.5)
    cv.line(left, y + 4, right, y + 4)
    y -= 14

    # Bottom stats
    cv.setFont("Helvetica", 7)
    cv.setFillColor(rl_colors.HexColor("#1A1A1A"))
    for stat_label, stat_key, fmt_str in [('Observations', 'nobs', '{:.0f}'),
                                            ('Adj. R-squared', 'r2_adj', '{:.4f}'),
                                            (f'Control Mean ({ctrl_label})', 'ctrl_mean', '{:.4f}')]:
        cv.drawString(left + 4, y - 4, stat_label)
        col_idx = 0
        for outcome in ['intl_mig', 'dom_mig']:
            for conf_var in ['mwar_own_any', 'cas_own_any']:
                key = f'{outcome}_{conf_var}_{suffix}'
                if key in results_dict:
                    val = results_dict[key].get(stat_key, np.nan)
                    if not np.isnan(val):
                        x = left + label_w + col_idx * data_w
                        cv.drawCentredString(x + data_w/2, y - 4, fmt_str.format(val))
                col_idx += 1
        y -= ROW_H

    y -= 10
    cv.setFont("Helvetica", 6)
    cv.setFillColor(NOTE_C)
    cv.drawString(left, y, "Intergenerational specification (Eq 8). Controls: District FE, mother's cohort FE, high caste, child girl, birth order FE.")
    cv.drawString(left, y - 10, "SE clustered at PSU level. *** p<0.01, ** p<0.05, * p<0.10")

# Create a new PDF with just Eq 8 tables, then merge
from pypdf import PdfReader, PdfWriter

eq8_pdf_path = f"{OUT_DIR}/phadera_eq8_tables.pdf"
page_size = landscape(letter)
pw, ph = page_size
c = pdf_canvas.Canvas(eq8_pdf_path, pagesize=page_size)

# Page 1: Table 6 - Eq 8 Phadera cohorts
draw_regression_pdf(c, 'Table 6: Intergenerational Impact (Eq 8, Phadera Cohorts)',
                     phadera_cohort_interactions, results_eq8, 'phadera_eq8', 'Mother 16-21', pw, ph)
c.showPage()

# Page 2: Table 7 - Eq 8 Our age groups
draw_regression_pdf(c, 'Table 7: Intergenerational Impact (Eq 8, Our Age Groups)',
                     our_cohort_interactions, results_eq8, 'agegroups_eq8', 'Mother 19-25', pw, ph)
c.showPage()
c.save()

# Merge with existing PDF
existing_pdf = f"{OUT_DIR}/phadera_replication_tables.pdf"
writer = PdfWriter()

reader1 = PdfReader(existing_pdf)
for page in reader1.pages:
    writer.add_page(page)

reader2 = PdfReader(eq8_pdf_path)
for page in reader2.pages:
    writer.add_page(page)

with open(existing_pdf, 'wb') as f:
    writer.write(f)

# Clean up temp PDF
os.remove(eq8_pdf_path)
print(f"  Updated PDF: {existing_pdf}")

print("\n" + "="*70)
print("  EQUATION 8 COMPLETE!")
print("="*70)
print(f"  Mother-child pairs used: {len(mc_sample)}")
print(f"  Results saved to: {OUT_DIR}/phadera_replication_tables.xlsx/.tex/.pdf")
